#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa dados do PEP RS, PMR-Outputs e PMR-Outcomes para o Supabase.

Uso:
  pip install openpyxl supabase
  python scripts/import_pep_supabase.py <planilha.xlsx> [--versao v2]

Variáveis de ambiente necessárias:
  SUPABASE_URL            https://dvqnlnxkwcrxbctujajl.supabase.co
  SUPABASE_SERVICE_ROLE_KEY  eyJ...
"""
import sys, os, argparse, math
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("Instale supabase-py: pip install supabase")
    sys.exit(1)

# ── Config ─────────────────────────────────────────────────────────────────

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://dvqnlnxkwcrxbctujajl.supabase.co"
)
SUPABASE_KEY = os.environ.get(
    "SUPABASE_SERVICE_ROLE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImR2cW5sbnhrd2NyeGJjdHVqYWpsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1MTU3Mzg5MiwiZXhwIjoyMDY3MTQ5ODkyfQ."
    "TTF0XttGd9Ytv11M3Bz8hcQ2T1MSx3gY-354bJhwWnc",
)

# ── Helpers ─────────────────────────────────────────────────────────────────

def val(cell) -> float:
    """Retorna valor numérico da célula (0 se vazio ou inválido)."""
    v = cell.value
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def txt(cell) -> str | None:
    """Retorna texto da célula (None se vazio)."""
    v = cell.value
    if v is None:
        return None
    return str(v).strip() or None

def safe_int(v) -> int | None:
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None

def flag(cell) -> int:
    """Retorna 0 ou 1 para células de entrega física (pode ser 1, 'x', True, etc.)."""
    v = cell.value
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return 1 if v else 0
    if isinstance(v, str):
        return 1 if v.strip().lower() in ('1', 'x', 'sim', 'yes', 'true') else 0
    return 0

def val_or_none(cell) -> float | None:
    """Retorna None se célula vazia, float se tiver valor."""
    v = cell.value
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    try:
        f = float(v)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None

# ── PEP RS ──────────────────────────────────────────────────────────────────

def import_pep(wb, versao: str, client) -> int:
    """Importa aba 'PEP RS' para tabela pep_entries."""
    if "PEP RS" not in wb.sheetnames:
        print("  ⚠  Aba 'PEP RS' não encontrada")
        return 0

    ws = wb["PEP RS"]
    rows = []

    for row_num in range(4, 247):  # L4-L245 dados, L246 total
        row = ws[row_num]
        ref = txt(row[0])  # col A
        if not ref or ref not in ("C", "SC", "P", "SP", "PT"):
            continue

        comp = safe_int(row[1].value)  # B
        prod = safe_int(row[2].value)  # C
        subp = safe_int(row[3].value)  # D
        pct  = safe_int(row[4].value)  # E
        wbs  = txt(row[8])             # I — código WBS (ex: "1.2.3.4")
        desc = txt(row[9])             # J — descrição autoritativa

        k_bid   = val_or_none(row[10])  # K — BRL BID
        l_local = val_or_none(row[11])  # L — BRL Local
        m_total = val_or_none(row[12])  # M — BRL Total
        n_atual = val(row[13])          # N — USD BID atual
        o_atual = val(row[14])          # O — USD Local atual
        p_atual = val(row[15])          # P — USD Total atual
        r_base  = val(row[17])          # R — USD BID arranque
        s_base  = val(row[18])          # S — USD Local arranque
        t_base  = val(row[19])          # T — USD Total arranque

        # Garantir que a linha tem colunas suficientes antes de ler
        ncols = len(row)
        pmr_ref   = txt(row[25]) if ncols > 25 else None   # Z
        pa_ref    = txt(row[26]) if ncols > 26 else None   # AA
        tipo_aq   = txt(row[27]) if ncols > 27 else None   # AB — Tipo Aquisição
        metodo_aq = txt(row[28]) if ncols > 28 else None   # AC — Método

        f2025 = flag(row[29]) if ncols > 29 else 0  # AD
        f2026 = flag(row[30]) if ncols > 30 else 0  # AE
        f2027 = flag(row[31]) if ncols > 31 else 0  # AF
        f2028 = flag(row[32]) if ncols > 32 else 0  # AG
        f2029 = flag(row[33]) if ncols > 33 else 0  # AH
        feop  = flag(row[34]) if ncols > 34 else 0  # AI

        d2025 = val_or_none(row[35]) if ncols > 35 else None  # AJ
        d2026 = val_or_none(row[36]) if ncols > 36 else None  # AK
        d2027 = val_or_none(row[37]) if ncols > 37 else None  # AL
        d2028 = val_or_none(row[38]) if ncols > 38 else None  # AM
        d2029 = val_or_none(row[39]) if ncols > 39 else None  # AN
        dtotal = val_or_none(row[40]) if ncols > 40 else None  # AO

        rows.append({
            "ref":               ref,
            "comp":              comp,
            "prod":              prod,
            "subp":              subp,
            "pct":               pct,
            "codigo_wbs":        wbs,
            "descricao":         desc,
            "k_reais_bid":       k_bid,
            "l_reais_local":     l_local,
            "m_reais_total":     m_total,
            "n_atual":           n_atual,
            "o_atual":           o_atual,
            "p_atual":           p_atual,
            "r_base":            r_base,
            "s_base":            s_base,
            "t_base":            t_base,
            "pmr_ref":           pmr_ref,
            "pa_ref":            pa_ref,
            "tipo_aquisicao":    tipo_aq,
            "metodo_aquisicao":  metodo_aq,
            "fisica_2025":       f2025,
            "fisica_2026":       f2026,
            "fisica_2027":       f2027,
            "fisica_2028":       f2028,
            "fisica_2029":       f2029,
            "fisica_eop":        feop,
            "desembolso_2025":   d2025,
            "desembolso_2026":   d2026,
            "desembolso_2027":   d2027,
            "desembolso_2028":   d2028,
            "desembolso_2029":   d2029,
            "desembolso_total":  dtotal,
            "versao":            versao,
            "linha_excel":       row_num,
        })

    if not rows:
        print("  ⚠  Nenhuma linha encontrada em PEP RS")
        return 0

    # Deletar versão anterior e re-importar
    client.table("pep_entries").delete().eq("versao", versao).execute()

    # Batch insert
    batch_size = 50
    for i in range(0, len(rows), batch_size):
        client.table("pep_entries").insert(rows[i : i + batch_size]).execute()

    return len(rows)

# ── PMR Outputs ─────────────────────────────────────────────────────────────

def _comp_from_codigo(codigo: str) -> str:
    """Deriva componente a partir do código do indicador (ex: '1.1' → 'C1')."""
    if not codigo:
        return ''
    c = str(codigo).strip()
    if c.startswith('R 1') or c.startswith('1'):
        return 'C1'
    if c.startswith('R 2') or c.startswith('2'):
        return 'C2'
    if c.startswith('3'):
        return 'C3'
    return 'Geral'


def import_pmr_outputs(wb, client) -> int:
    """Importa aba 'PMR-Outputs' para tabela pmr_outputs.

    Estrutura da planilha (20260306_Revisão_PEP_PMR_V2_PN.xlsx):
      Cabeçalho: linha 4
      Dados: linha 5 em diante
      B(1)=codigo  C(2)=descricao  D(3)=unidade  E(4)=linha_base
      H(7)=meta_ano1  M(12)=meta_fim_projeto  N(13)=meio_verificacao
    """
    aba = None
    for name in wb.sheetnames:
        if "output" in name.lower():
            aba = wb[name]
            break
    if not aba:
        print("  ⚠  Aba PMR-Outputs não encontrada")
        return 0

    rows = []
    for row in aba.iter_rows(min_row=5):           # dados a partir da linha 5
        codigo = txt(row[1])
        if not codigo:
            continue
        # Pular linhas de cabeçalho de componente (código sem ponto, ex: '1', '2', '3')
        if codigo.isdigit():
            continue
        descricao = txt(row[2])
        if not descricao:
            continue
        meta = val(row[12]) or None                # col M = meta fim do projeto
        meta_periodo = val(row[7]) or None          # col H = meta ano 1
        realizado = 0.0                             # campo preenchido manualmente
        pct = 0.0
        rows.append({
            "componente":    _comp_from_codigo(codigo),
            "produto":       descricao,
            "codigo":        codigo,
            "descricao":     descricao,
            "unidade":       txt(row[3]),           # col D
            "linha_base":    val(row[4]) or None,  # col E
            "meta_contrato": meta,
            "meta_periodo":  meta_periodo,
            "realizado":     realizado,
            "pct_realizado": pct,
        })

    if not rows:
        return 0

    client.table("pmr_outputs").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    for i in range(0, len(rows), 50):
        client.table("pmr_outputs").insert(rows[i : i + 50]).execute()
    return len(rows)

# ── PMR Outcomes ─────────────────────────────────────────────────────────────

def import_pmr_outcomes(wb, client) -> int:
    """Importa aba 'PMR-Outcomes' para tabela pmr_outcomes.

    Estrutura da planilha (20260306_Revisão_PEP_PMR_V2_PN.xlsx):
      Cabeçalho: linha 4
      Dados: linha 6 em diante (linha 5 = Objetivo Geral sem código)
      B(1)=codigo  D(3)=descricao(PT)  E(4)=unidade  F(5)=linha_base
      H(7)=meta_ano1  M(12)=meta_fim_projeto  N(13)=fonte_dados
    """
    aba = None
    for name in wb.sheetnames:
        if "outcome" in name.lower():
            aba = wb[name]
            break
    if not aba:
        print("  ⚠  Aba PMR-Outcomes não encontrada")
        return 0

    rows = []
    for row in aba.iter_rows(min_row=6):           # dados a partir da linha 6
        codigo = txt(row[1])
        if not codigo:
            continue
        descricao = txt(row[3])                     # col D = versão em português
        if not descricao:
            descricao = txt(row[2])                 # fallback: versão em espanhol (col C)
        if not descricao:
            continue
        meta_oc = val(row[12]) or None              # col M = meta fim do projeto
        linha_base_raw = row[5].value if len(row) > 5 else None
        # linha_base pode ser número ou string como 'LB[1]'
        linha_base = None
        if isinstance(linha_base_raw, (int, float)):
            linha_base = float(linha_base_raw)
        rows.append({
            "componente":    _comp_from_codigo(codigo),
            "objetivo":      descricao,
            "codigo":        codigo,
            "descricao":     descricao,
            "unidade":       txt(row[4]),           # col E
            "linha_base":    linha_base,
            "meta_contrato": meta_oc,
            "realizado":     0.0,
            "pct_realizado": 0.0,
            "fonte_dados":   txt(row[13]) if len(row) > 13 else None,  # col N
        })

    if not rows:
        return 0

    client.table("pmr_outcomes").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    for i in range(0, len(rows), 50):
        client.table("pmr_outcomes").insert(rows[i : i + 50]).execute()
    return len(rows)

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Importa PEP/PMR para Supabase")
    parser.add_argument("planilha", help="Caminho para o arquivo .xlsx")
    parser.add_argument("--versao", default="v1", help="Versão do PEP (default: v1)")
    args = parser.parse_args()

    planilha = Path(args.planilha)
    if not planilha.exists():
        print(f"Arquivo não encontrado: {planilha}")
        sys.exit(1)

    print(f"\n📊 Importando: {planilha.name}")
    print(f"   Versão PEP: {args.versao}")
    print(f"   Supabase: {SUPABASE_URL}\n")

    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    wb = openpyxl.load_workbook(planilha, read_only=True, data_only=True)
    print(f"   Abas encontradas: {', '.join(wb.sheetnames)}\n")

    n_pep = import_pep(wb, args.versao, client)
    print(f"   ✓ PEP RS: {n_pep} linhas importadas")

    n_out = import_pmr_outputs(wb, client)
    print(f"   ✓ PMR-Outputs: {n_out} indicadores importados")

    n_ouc = import_pmr_outcomes(wb, client)
    print(f"   ✓ PMR-Outcomes: {n_ouc} indicadores importados")

    # Registrar no sync_log
    try:
        client.table("sync_log").insert([
            {
                "tabela_destino": "pep_entries",
                "fonte": f"pep_xlsx_expanded:{planilha.name}",
                "versao": args.versao,
                "registros_lidos": n_pep,
                "registros_inseridos": n_pep,
                "registros_atualizados": 0,
                "registros_erro": 0,
                "status": "ok" if n_pep > 0 else "parcial",
                "executado_por": "import_pep_supabase",
            },
            {
                "tabela_destino": "pmr_outputs",
                "fonte": f"pep_xlsx_expanded:{planilha.name}",
                "versao": args.versao,
                "registros_lidos": n_out,
                "registros_inseridos": n_out,
                "registros_atualizados": 0,
                "registros_erro": 0,
                "status": "ok" if n_out > 0 else "parcial",
                "executado_por": "import_pep_supabase",
            },
            {
                "tabela_destino": "pmr_outcomes",
                "fonte": f"pep_xlsx_expanded:{planilha.name}",
                "versao": args.versao,
                "registros_lidos": n_ouc,
                "registros_inseridos": n_ouc,
                "registros_atualizados": 0,
                "registros_erro": 0,
                "status": "ok" if n_ouc > 0 else "parcial",
                "executado_por": "import_pep_supabase",
            },
        ]).execute()
        print("   ✓ sync_log atualizado")
    except Exception as e:
        print(f"   ⚠  sync_log falhou: {e}")

    print(f"\n✅ Importação concluída em {datetime.now().strftime('%H:%M:%S')}")
    print(f"   Total: {n_pep + n_out + n_ouc} registros\n")

if __name__ == "__main__":
    main()
